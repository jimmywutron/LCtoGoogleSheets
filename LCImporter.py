#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Dec  4 20:29:07 2024

@author: jimmywu
"""
#%% Store LeadershipConnect data in a dictionary

from tkinter import Tk
from tkinter.filedialog import askopenfilename
import openpyxl, pprint, os
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font
import pandas as pd
import ezsheets
import LCData

#%% Ask for LC File and Verify that it is an excel file

Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file

try:
    if filename.endswith('.csv'):
        # Reading the csv file
        df_new = pd.read_csv(filename)
        # saving xlsx file
        filenameshort = filename[:-3]
        filename = filenameshort + 'xlsx'
        GFG = df_new.to_excel(filename, index=False)
    wb = openpyxl.load_workbook(filename)
    print(f'Success! The file "{filename}" was created and opened.')
except:
    print("Invalid File Type. Please ensure that the file selected is an XLSM file")


#%%

govTypeKey = input("Are you updating Congressional or Executive branch data? For Congressional type '1'. For Executive, type '2': ")

if govTypeKey == '1':
    govType = 'Congressional'
    print(govType)
elif govTypeKey == '2':
    govType = 'Executive'
    print(govType)
else:
    print("Please input a valid number")

#%% Read and save person data as a dictionary
dateString = str(datetime.utcnow())

print("Opening workbook...")
# Create list of data categories
# print(wb.sheetnames)
sheet = wb[str(wb.sheetnames[0])]
categories = []
print("Creating list of categories")
for col in range(1, sheet.max_column+1):
    #print("Adding " + sheet.cell(1,col).value + "to the cateogry list")
    categories.append(sheet.cell(1,col).value)
print("Added %s categories" %(str(len(categories))))
    #print(sheet[str(col)+'1'].value)
    #print('added %s to the list of categories') % (sheet[str(col)+'1'].value)
people = {}
peopleCount = 0
# Read each row's data
for row in range(2, sheet.max_row+1):
    # Add people data to a dictionary, with the leadership ID number as the main key 
    LCid = str(sheet.cell(row, 1).value)
    people.setdefault(LCid, {})
    for cat in range(2, len(categories)+1):
        people[LCid][str(sheet.cell(1,cat).value)] = str(sheet.cell(row, cat).value)
    people[LCid]["govType"] = govType
    people[LCid]["Last Updated"] = dateString
    office = ""
    if people[LCid]['Organization Name (Parent)'] != 'None':
        office += people[LCid]['Organization Name (Parent)']
    if people[LCid]['Organization Name (Intermediate)'] != 'None':
        if office == "":
            office += people[LCid]['Organization Name (Intermediate)']
        else:   
            office += ", " + people[LCid]['Organization Name (Intermediate)']
    if people[LCid]['Organization Name (Child)'] != 'None':
        if office == "":
            office += people[LCid]['Organization Name (Child)']
        else:
            office += ", " + people[LCid]['Organization Name (Child)']
    if office != "":
        people[LCid]['Current Office'] = office
    else:
        people[LCid]['Current Office'] = "None"
    
    peopleCount += 1
    
print("Added %s people" %(peopleCount))

#%% Prompt user to add another spreadsheet

continueKey = input("Would you like to add another spreadsheet? Type '1' for yes, '2' for no： ")
if continueKey == '1':
    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file

    try:
        if filename.endswith('.csv'):
            # Reading the csv file
            df_new = pd.read_csv(filename)
            # saving xlsx file
            filenameshort = filename[:-3]
            filename = filenameshort + 'xlsx'
            GFG = df_new.to_excel(filename, index=False)
        wb = openpyxl.load_workbook(filename)
        print(f'Success! The file "{filename}" was created and opened.')
    except:
        print("Invalid File Type. Please ensure that the file selected is an XLSM file")
        
    govTypeKey = input("Are you updating Congressional or Executive branch data? For Congressional type '1'. For Executive, type '2': ")
    if govTypeKey == '1':
        govType = 'Congressional'
        print(govType)
    elif govTypeKey == '2':
        govType = 'Executive'
        print(govType)
    else:
        print("Please input a valid number")
        
        
    dateString = str(datetime.utcnow())

    print("Opening workbook...")
    peopleCount = 0
    sheet = wb[str(wb.sheetnames[0])]
    # Read each row's data
    for row in range(2, sheet.max_row+1):
        # Add people data to a dictionary, with the leadership ID number as the main key 
        LCid = str(sheet.cell(row, 1).value)
        people.setdefault(LCid, {})
        for cat in range(2, len(categories)+1):
            people[LCid][str(sheet.cell(1,cat).value)] = str(sheet.cell(row, cat).value)
        people[LCid]["govType"] = govType
        people[LCid]["Last Updated"] = dateString
        office = ""
        if people[LCid]['Organization Name (Parent)'] != 'None':
            office += people[LCid]['Organization Name (Parent)']
        if people[LCid]['Organization Name (Intermediate)'] != 'None':
            if office == "":
                office += people[LCid]['Organization Name (Intermediate)']
            else:   
                office += ", " + people[LCid]['Organization Name (Intermediate)']
        if people[LCid]['Organization Name (Child)'] != 'None':
            if office == "":
                office += people[LCid]['Organization Name (Child)']
            else:
                office += ", " + people[LCid]['Organization Name (Child)']
        if office != "":
            people[LCid]['Current Office'] = office
        else:
            people[LCid]['Current Office'] = "None"
        
        peopleCount += 1
        
    print("Added %s people" %(peopleCount))



#%%
keys = people.keys()
print("Number of people: " + str(len(keys)))
keyList = []
for key in keys:
    keyList.append(key)

#%% Writing results as a .py file
print('Writing results......')
resultFile = open('LCData.py', 'w')
resultFile.write('allData = ' + pprint.pformat(people))
resultFile.close()
print('Done.')

#%% Update Google Sheets with LCData

# Initialize sheet + variables

people = LCData.allData
spreadsheetID = '1qB0typdQr1e3KK38lp9e8zQi1kMf2Zn2Ud30Gakya58'
ss = ezsheets.Spreadsheet(spreadsheetID)
ss.title
sheet = ss[0]
headings = sheet.getRow(1)
LeadershipConnectIDCol = headings.index('LeadershipConnectID')+1
maxCol = len(headings)
maxRow = len(sheet.getColumn(1))
print("ID is in column %s, titled %s" %(LeadershipConnectIDCol,sheet[LeadershipConnectIDCol,1]))
print("Updating %s" %(ss.title))
# Start of LeadershipConnect Fields
lastUpdatedCol = headings.index('Last Updated (LC)')+1


#%%

ss.refresh()
updateCount = 0
for row in range(1000,maxRow):
    ID = sheet[LeadershipConnectIDCol,row]
    print("ID: %s" %(ID))
    if ID in people.keys():
        print("Updating ID: %s" %(ID))
        updateCount += 1
        for col in range(lastUpdatedCol,maxCol+1):
            key = sheet[col,1][:-5]
            if key in people[ID].keys():
                print("Updating %s" %(key))
                print("New Value: %s" %(people[ID][key]))
# =============================================================================
#                 if people[ID][key] == 'None':
#                     sheet[col,row] = ""
#                 else:
# =============================================================================
                sheet[col,row] = people[ID][key]
print("Finished updating %s entries" %(updateCount))     


#%%
# =============================================================================
# import LCData
# for key in keyList:
#     print(key + ": \n")
#     for cat in range(1,20):
#         print(categories[cat] + ": " + LCData.allData[key][categories[cat]])
# len(LCData.allData)
# =============================================================================
        
#%% Use LCData to udpate an excel sheet

# =============================================================================
# CATEGORIES_TO_WRITE = ["Prefix","First Name","Middle Name", "Nickname", "Last Name", "Maiden Name", "Suffix", "Credentials", "Title", "Member Name", "Organization Sector", "Organization Name (Parent)" ]
# toWrite = openpyxl.load_workbook("/Users/jimmywu/Downloads/20241210_TechLabs_AlumMasterList.xlsx")
# sheetToWrite = toWrite.active
# 
# # Write column titles into existing sheet
# print(sheetToWrite.max_column)
# col = sheetToWrite.max_column
# for cat in categories:
#     sheetToWrite.cell(2, col).value = cat
#     sheetToWrite.cell(2, col).font = Font(name='Times New Roman', bold=True)
#     col += 1
# toWrite.save("/Users/jimmywu/Downloads/20241210_TechLabs_AlumMasterList_Updated.xlsx")
# 
# # Fill out data with LC Data
# idCol = column_index_from_string('E')
# for row in range(2, sheetToWrite.max_row):
#     LCID = str(sheetToWrite.cell(row, idCol).value)
#     print(LCID)
#     if LCID in people.keys():
#         # use the heading column as the key
#         print("Updating data for PersonID: %s" %(LCID))
#         for col in range(32, 32 + len(people[LCID].keys())):
#             cat = sheetToWrite.cell(2, col).value
#             if cat in people[LCID].keys():
#                 print("Updating category: %s" % cat)
#                 if people[LCID][cat] == "None":
#                         sheetToWrite.cell(row, col).value = ""
#                 else:
#                     sheetToWrite.cell(row, col).value = people[LCID][cat]
#     
#             
# 
# toWrite.save("/Users/jimmywu/Downloads/20241210_TechLabs_AlumMasterList_Updated.xlsx")
# 
# =============================================================================
# =============================================================================
# #%%
# 
# people.keys()
# for row in range(2, sheetToWrite.max_row):
#     LCID = sheetToWrite.cell(row, idCol).value
#     print(LCID)
#     print(str(LCID) in people.keys())
# 
# "1774560" in people.keys()
# print(cat)
# print(people["1774560"].keys())
# "Member Committees" in people["1774560"].keys()
# cat in people["1774560"].keys()
# 
# people["1774560"][ 'Member Committees']
# cat
# 
# =============================================================================

